#!/usr/bin/env python3

import sys
import platform
import signal
from contextlib import contextmanager
import threading
import queue
from functools import wraps
import os

# === Configuration Settings ===
# Set the directory path - Configure this to point to your Minecraft directory
#directory_path = os.path.normpath('C:/Users/juke32/AppData/Roaming/.minecraft')
#directory_path = os.path.normpath('D:/dump/randomdrive')
directory_path = os.path.normpath('D:/dump')

# Timeout settings - Adjust these to control processing timeouts
BASE_TIMEOUT = 1  # Base timeout in seconds for file processing
MAX_TIMEOUT = 4  # Maximum timeout in seconds for any file
SIZE_TIMEOUT_RATIO = 1024 * 1024  # Add 1 second for each MB of file size

def truncate(text, length=32):
    """Truncate text to specified length"""
    if len(text) <= length:
        return text
    return text[:length-3] + "..."

print("\n=== Python Environment Info ===")
print(truncate(f"Python: {sys.version.split()[0]}"))
print(truncate(f"System: {platform.system()}"))
print("=============================\n")

try:
    print("Importing nbtlib...")
    import nbtlib
    print("Import successful")
except ImportError as e:
    print("\nERROR: Missing nbtlib package")
    print("Run: pip install nbtlib")
    input("\nPress Enter to exit...")
    sys.exit(1)

import gzip
import re
import traceback
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# Patterns for finding seeds in logs
seed_patterns = [
    re.compile(r'(?:seed|Seed)[:|\s]+(-?\d{1,19})'),
    re.compile(r'/seed\s+(-?\d{1,19})'),
    re.compile(r'Seed:\s*\[(-?\d{1,19})\]'),
    re.compile(r'RandomSeed:\s*(-?\d{1,19})'),
    re.compile(r'worldGenSeed:\s*(-?\d{1,19})'),
    re.compile(r'World\s+Seed:\s*(-?\d{1,19})'),
    re.compile(r'Random\s+seed:\s*(-?\d{1,19})'),
    re.compile(r'Seed\s+used:\s*(-?\d{1,19})')
]

# Version patterns
version_patterns = [
    re.compile(r'(?:Minecraft|MC)\s*(?:version|v\.?|:)\s*([\d\.]+(?:-pre\d+)?(?:\w+)?)', re.IGNORECASE),
    re.compile(r'(?:Data|Version)\s*(?:version|v\.?|:)\s*([\d\.]+(?:-pre\d+)?(?:\w+)?)', re.IGNORECASE)
]

# Game mode patterns
gamemode_patterns = [
    re.compile(r'(?:gamemode|GameType)[:\s]+(survival|creative|adventure|spectator)', re.IGNORECASE),
    re.compile(r'/gamemode\s+(survival|creative|adventure|spectator)', re.IGNORECASE)
]

# Search terms to include in log results
search_terms = [
    'seed', 'Seed', '/seed', 'minecraft', 'world', 'generate', 'creating',
    'version', 'gamemode', 'GameType', 'difficulty', 'hardcore',
    'cheats', 'allowCommands', 'DataVersion', 'WanderingTrader',
    'SpawnX', 'SpawnY', 'SpawnZ', 'Time', 'LastPlayed', 'SizeOnDisk',
    'Starting minecraft server', 'Loading world', 'Preparing start region',
    'World Settings', 'World Generation', 'Random seed', 'World seed'
]

# Example of ignored seeds
ignored_seeds = [
    "1234567890",
    "9876543210"
]

# Global variables for Excel workbook
wb = None
ws_data = None
ws_errors = None
ws_log = None
ws_all_seeds = None
ws_corrupted = None
ws_potential = None  # Added potential seeds worksheet
error_fill = None
unique_seeds = {}
potential_seeds = {}  # Track potential seeds and their contexts

# Initialize row counters
row_data = 2
row_errors = 2
row_log = 2
row_all_seeds = 2
row_corrupted = 2

# Initialize counters for processed and saved files
processed_files = 0
saved_entries = 0
errors_encountered = 0
corrupted_files = 0

class TimeoutException(Exception):
    pass

def timeout_handler(seconds):
    """Cross-platform timeout decorator that works on both Windows and Unix"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            result = queue.Queue()
            
            def worker():
                try:
                    result.put(('success', func(*args, **kwargs)))
                except Exception as e:
                    result.put(('error', e))
            
            thread = threading.Thread(target=worker)
            thread.daemon = True
            thread.start()
            thread.join(seconds)
            
            if thread.is_alive():
                raise TimeoutException("Operation timed out")
            
            status, value = result.get()
            if status == 'error':
                raise value
            return value
        return wrapper
    return decorator

def initialize_excel_workbook():
    """Initialize Excel workbook and worksheets"""
    global wb, ws_data, ws_errors, ws_log, ws_all_seeds, ws_corrupted, ws_potential, error_fill
    
    # Create workbook and sheets
    wb = Workbook()
    ws_all_seeds = wb.active  # First sheet
    ws_all_seeds.title = "All Seeds"
    ws_log = wb.create_sheet(title="Log Results")
    ws_data = wb.create_sheet(title="Data")
    ws_errors = wb.create_sheet(title="Errors")
    ws_corrupted = wb.create_sheet(title="Corrupted Files")
    ws_potential = wb.create_sheet(title="Random Strings")  # Renamed from "Potential Seeds"
    
    # Create highlight fill for errors
    error_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    
    # Set headers and adjust column widths
    setup_excel_headers()
    adjust_column_widths()

def setup_excel_headers():
    """Set up headers for all Excel sheets"""
    # All Seeds tab headers (now first)
    ws_all_seeds['A1'] = 'Seed Value'
    ws_all_seeds['B1'] = 'First Found In'
    ws_all_seeds['C1'] = 'World Name'
    ws_all_seeds['D1'] = 'Game Mode'
    ws_all_seeds['E1'] = 'Generator'
    ws_all_seeds['F1'] = 'Version'
    ws_all_seeds['G1'] = 'Last Played'
    ws_all_seeds['H1'] = 'Path'
    ws_all_seeds['I1'] = 'Times Found'
    ws_all_seeds['J1'] = 'Total Time'
    ws_all_seeds['K1'] = 'Spawn Location'
    ws_all_seeds['L1'] = 'Data Version'
    ws_all_seeds['M1'] = 'Difficulty'
    ws_all_seeds['N1'] = 'Hardcore'
    ws_all_seeds['O1'] = 'Allow Commands'
    ws_all_seeds['P1'] = 'Size on Disk'
    
    # Log Results tab headers (second)
    ws_log['A1'] = 'File Name'
    ws_log['B1'] = 'Path'
    ws_log['C1'] = 'Log Line'
    ws_log['D1'] = 'Extracted Seed'
    
    # Data tab headers (third)
    ws_data['A1'] = 'File Name'
    ws_data['B1'] = 'Random Seed'
    ws_data['C1'] = 'Total Time'
    ws_data['D1'] = 'Generator Name'
    ws_data['E1'] = 'Level Name'
    ws_data['F1'] = 'Game Mode'
    ws_data['G1'] = 'Spawn Location'
    ws_data['H1'] = 'Path'
    ws_data['I1'] = 'Version'
    ws_data['J1'] = 'Data Version'
    ws_data['K1'] = 'Last Played'
    ws_data['L1'] = 'Size on Disk'
    ws_data['M1'] = 'Difficulty'
    ws_data['N1'] = 'Hardcore'
    ws_data['O1'] = 'Allow Commands'
    ws_data['P1'] = 'Has Errors'
    
    # Errors tab headers (fourth)
    ws_errors['A1'] = 'File Name'
    ws_errors['B1'] = 'Error Message'
    ws_errors['C1'] = 'Path'
    ws_errors['D1'] = 'Traceback'

    # Corrupted Files tab headers (fifth)
    ws_corrupted['A1'] = 'File Name'
    ws_corrupted['B1'] = 'Path'
    ws_corrupted['C1'] = 'Partial Data Retrieved'
    ws_corrupted['D1'] = 'Error Details'

    # Random Strings tab headers (last, renamed from Potential Seeds)
    ws_potential['A1'] = 'Number'
    ws_potential['B1'] = 'Found In'
    ws_potential['C1'] = 'Context'
    ws_potential['D1'] = 'Line'
    ws_potential['E1'] = 'Path'
    ws_potential['F1'] = 'Confidence'

def adjust_column_widths():
    """Adjust column widths for all Excel sheets"""
    # Data tab column widths
    ws_data.column_dimensions['A'].width = 20
    ws_data.column_dimensions['B'].width = 40
    ws_data.column_dimensions['C'].width = 15
    ws_data.column_dimensions['D'].width = 15
    ws_data.column_dimensions['E'].width = 10
    ws_data.column_dimensions['F'].width = 15
    ws_data.column_dimensions['G'].width = 20
    ws_data.column_dimensions['H'].width = 50
    ws_data.column_dimensions['I'].width = 10
    ws_data.column_dimensions['J'].width = 15
    ws_data.column_dimensions['K'].width = 20
    ws_data.column_dimensions['L'].width = 15
    ws_data.column_dimensions['M'].width = 15
    ws_data.column_dimensions['N'].width = 10
    ws_data.column_dimensions['O'].width = 15
    ws_data.column_dimensions['P'].width = 15
    
    # Errors tab column widths
    ws_errors.column_dimensions['A'].width = 20
    ws_errors.column_dimensions['B'].width = 50
    ws_errors.column_dimensions['C'].width = 50
    ws_errors.column_dimensions['D'].width = 100

    # Log Results tab column widths
    ws_log.column_dimensions['A'].width = 20
    ws_log.column_dimensions['B'].width = 50
    ws_log.column_dimensions['C'].width = 100
    ws_log.column_dimensions['D'].width = 20

    # All Seeds tab column widths
    ws_all_seeds.column_dimensions['A'].width = 30  # Seed Value
    ws_all_seeds.column_dimensions['B'].width = 20  # First Found In
    ws_all_seeds.column_dimensions['C'].width = 30  # World Name
    ws_all_seeds.column_dimensions['D'].width = 10  # Game Mode
    ws_all_seeds.column_dimensions['E'].width = 15  # Generator
    ws_all_seeds.column_dimensions['F'].width = 15  # Version
    ws_all_seeds.column_dimensions['G'].width = 20  # Last Played
    ws_all_seeds.column_dimensions['H'].width = 50  # Path
    ws_all_seeds.column_dimensions['I'].width = 10  # Times Found
    ws_all_seeds.column_dimensions['J'].width = 20  # Total Time
    ws_all_seeds.column_dimensions['K'].width = 30  # Spawn Location
    ws_all_seeds.column_dimensions['L'].width = 15  # Data Version
    ws_all_seeds.column_dimensions['M'].width = 15  # Difficulty
    ws_all_seeds.column_dimensions['N'].width = 10  # Hardcore
    ws_all_seeds.column_dimensions['O'].width = 10 # Allow Commands
    ws_all_seeds.column_dimensions['P'].width = 15  # Size on Disk
    
    # Corrupted Files tab column widths
    ws_corrupted.column_dimensions['A'].width = 20
    ws_corrupted.column_dimensions['B'].width = 50
    ws_corrupted.column_dimensions['C'].width = 20
    ws_corrupted.column_dimensions['D'].width = 30

    # Random Strings tab column widths
    ws_potential.column_dimensions['A'].width = 25  # Number
    ws_potential.column_dimensions['B'].width = 30  # Found In
    ws_potential.column_dimensions['C'].width = 40  # Context
    ws_potential.column_dimensions['D'].width = 100  # Line
    ws_potential.column_dimensions['E'].width = 50  # Path
    ws_potential.column_dimensions['F'].width = 15  # Confidence

def print_debug_info():
    """Print system and environment information for debugging"""
    print("\n=== Debug Information ===")
    print(f"Python version: {sys.version}")
    print(f"Platform: {platform.platform()}")
    print(f"System: {platform.system()}")
    print(f"Directory path: {directory_path}")
    print("========================\n")

def is_completely_empty(file_path):
    """Check if a file is completely empty (no data at all)"""
    try:
        if os.path.getsize(file_path) == 0:
            return True
        with open(file_path, 'rb') as f:
            content = f.read(1024)
            return len(content.strip()) == 0
    except Exception:
        return False

def find_seed_in_nbt(nbt_data):
    """Recursively search for seed values in NBT data"""
    if isinstance(nbt_data, dict):
        if 'RandomSeed' in nbt_data:
            return str(nbt_data['RandomSeed'])
        if 'seed' in nbt_data:
            return str(nbt_data['seed'])
        
        if 'WorldGenSettings' in nbt_data:
            if isinstance(nbt_data['WorldGenSettings'], dict):
                if 'seed' in nbt_data['WorldGenSettings']:
                    return str(nbt_data['WorldGenSettings']['seed'])
                
        if 'DimensionData' in nbt_data:
            for dim_data in nbt_data['DimensionData'].values():
                seed = find_seed_in_nbt(dim_data)
                if seed is not None:
                    return seed
        
        for key, value in nbt_data.items():
            if isinstance(value, dict):
                result = find_seed_in_nbt(value)
                if result is not None:
                    return result
            elif isinstance(value, list):
                for item in value:
                    if isinstance(item, dict):
                        result = find_seed_in_nbt(item)
                        if result is not None:
                            return result
    return None

def is_valid_seed(seed_str):
    """Check if a string could be a valid Minecraft seed"""
    if not seed_str:
        return False
    seed_str = seed_str.strip()
    if re.match(r'^-?\d{1,19}$', seed_str):
        return True
    if len(seed_str) <= 50:
        return True
    return False

def get_generator_name(nbt_data):
    """Get the generator name from NBT data"""
    try:
        if 'WorldGenSettings' in nbt_data:
            settings = nbt_data['WorldGenSettings']
            if isinstance(settings, dict):
                return settings.get('bonus_chest', 'default')
        return nbt_data.get('generatorName', 'default')
    except:
        return 'Unknown'

def is_binary_content(content):
    """Less aggressive check if content appears to be binary data"""
    try:
        # Convert string to bytes if needed
        if isinstance(content, str):
            content = content.encode('utf-8')
        
        # Check first chunk for binary characters
        chunk = content[:1024] if len(content) > 1024 else content
        text_characters = bytes(range(32, 127)) + b'\n\r\t\f\b'
        binary_chars = bytes(set(range(256)) - set(text_characters))
        
        # Count null bytes and other binary characters
        null_count = chunk.count(b'\x00')
        binary_count = len([x for x in chunk if x in binary_chars])
        
        # If more than 50% nulls or 70% binary, consider it binary
        if len(chunk) == 0:
            return False
        
        null_ratio = null_count / len(chunk)
        binary_ratio = binary_count / len(chunk)
        
        return null_ratio > 0.5 or binary_ratio > 0.7
    except:
        return True  # If any error occurs, assume it's binary

def is_meaningful_log(line):
    """Check if a log line contains meaningful information we want to track"""
    # Skip common unnecessary lines
    skip_patterns = [
        'RCON running on',
        'Starting minecraft server version',
        'Loading properties',
        'Default game type:',
        'Preparing level "',
        'Preparing start region',
        'Time elapsed:',
        'Done (',
        '[Server thread/INFO]',
        'Starting Minecraft server on',
        'Using epoll channel type',
        'Preparing spawn area:',
        'Starting GS4 status listener',
        'Thread RCON Listener started',
        'RCON running on',
        '[Server Shutdown Thread/INFO]',
        '[Server thread/WARN]',
        '[Server thread/ERROR]',
        'Stopping server',
        'Saving players',
        'Saving worlds',
        'Saving chunks',
        'ThreadedAnvilChunkStorage',
        'Connection #',
        'UUID of player',
        'logged in with entity id',
        'Disconnecting',
        'lost connection:',
        'left the game',
        'joined the game',
        '[Not Secure]',
        '[Async Chat Thread',
        '[User Authenticator',
        'moved wrongly!',
        'moved too quickly!',
        'moved too far!',
        'Playing effect',
        'Particle',
        'Saving crash report',
        'Stopping the server',
        'Commencing server shutdown',
        'Saving chunks for level',
        'Starting integrated minecraft server',
        'Changing view distance to',
        'Preparing dimension',
        'Loaded',
        'Generated new',
        'logged in successfully',
        'moved too quickly',
        'moved wrongly',
        'tried command',
        'issued server command',
        'Fetching addPacket',
        'handleDisconnection',
        'Reached end of stream',
        'closed connection',
        'disconnected',
        'Disconnecting',
        'Stopping singleplayer server',
        'Stopping server',
        'Server thread/INFO',
        'Starting integrated server',
        'Saving and pausing game',
        'Saving the game',
        'Saving chunks for level',
        'Saving',
        'Starting Server',
        'Loading dimension',
        'Preparing spawn area',
        'Preparing spawn region',
        'Preparing level',
        'Preparing start region',
        'Time elapsed',
        'Done',
        'For help',
        'Unknown command',
        'Invalid command syntax',
        'Kicked',
        'banned',
        'unbanned',
        'op',
        'deop',
        'whitelist',
        'teleport',
        'gamemode',
        'difficulty',
        'time set',
        'weather',
        'xp',
        'give',
        'kill',
        'scoreboard',
        'advancement',
        'recipe',
        'function',
        'debug',
        'reload',
        'save-all',
        'save-off',
        'save-on',
        'stop',
        'tell',
        'msg',
        'w',
        'me',
        'say',
        'trigger',
        'worldborder',
        'spawnpoint',
        'setworldspawn',
        'gamerule',
        'title',
        'particle',
        'playsound',
        'stopsound',
        'worldborder',
        'defaultgamemode',
        'enchant',
        'experience',
        'fill',
        'setblock',
        'summon',
        'tp',
        'spreadplayers',
        'achievement',
        'clear',
        'effect',
        'replaceitem',
        'stats',
        'testfor',
        'toggledownfall',
        'weather',
        'xp'
    ]
    
    # Skip if line starts with any of these patterns
    if any(line.startswith(pattern) for pattern in skip_patterns):
        return False
        
    # Skip if line contains any of these patterns
    if any(pattern in line for pattern in skip_patterns):
        return False
    
    # Only keep lines that might have useful information
    keep_patterns = [
        'seed',
        'Seed',
        '/seed',
        'minecraft',
        'world',
        'generate',
        'creating',
        'version',
        'gamemode',
        'GameType',
        'difficulty',
        'hardcore',
        'cheats',
        'allowCommands',
        'DataVersion',
        'WanderingTrader',
        'SpawnX',
        'SpawnY',
        'SpawnZ',
        'Time',
        'LastPlayed',
        'SizeOnDisk',
        'World Settings',
        'World Generation',
        'Random seed',
        'World seed'
    ]
    
    return any(pattern in line for pattern in keep_patterns)

def is_potential_seed(text):
    """Check if a string might be a seed based on expanded criteria"""
    # Must be numeric and reasonable length for a seed
    if not text or len(text) > 20:
        return False
    
    # Check if it's purely numeric (positive or negative)
    if re.match(r'^-?\d+$', text):
        return True
    
    # Check if it looks like a seed in scientific notation
    if re.match(r'^-?\d+\.?\d*[eE][+-]?\d+$', text):
        return True
        
    return False

def find_potential_seeds(line, filename, root):
    """Find potential seeds in a line of text"""
    global ws_potential
    
    # Skip if line is too short or doesn't contain numbers
    if len(line) < 3 or not any(c.isdigit() for c in line):
        return
    
    # Look for numbers in the context of world generation or similar
    seed_contexts = [
        'world', 'gen', 'seed', 'random', 'create', 'new', 'generate',
        'level', 'map', 'terrain', 'dimension', 'spawn'
    ]
    
    # Split line into words and examine each one
    words = line.split()
    for i, word in enumerate(words):
        if is_potential_seed(word):
            # Get surrounding context (up to 3 words before and after)
            start = max(0, i - 3)
            end = min(len(words), i + 4)
            context = ' '.join(words[start:end])
            
            # Calculate confidence based on context
            confidence = 'Low'
            if any(term.lower() in line.lower() for term in seed_contexts):
                confidence = 'Medium'
                if any(term.lower() in line.lower() for term in ['seed', 'world seed', 'random seed']):
                    confidence = 'High'
            
            # Store in potential seeds dictionary
            if word not in potential_seeds:
                potential_seeds[word] = {
                    'filename': filename,
                    'context': context,
                    'line': line,
                    'path': root,
                    'confidence': confidence
                }

def get_timeout_for_size(file_path):
    """Calculate appropriate timeout based on file size"""
    try:
        size = os.path.getsize(file_path)
        # Base timeout plus additional time based on file size
        timeout = BASE_TIMEOUT + (size / SIZE_TIMEOUT_RATIO)
        # Cap at MAX_TIMEOUT to prevent extremely long waits
        return min(MAX_TIMEOUT, max(BASE_TIMEOUT, timeout))
    except:
        return BASE_TIMEOUT  # Default to base timeout if can't determine size

def process_log_content(log_data, filename, root):
    """Process log content for seed information"""
    global row_log, row_all_seeds, ws_log, ws_all_seeds, unique_seeds, potential_seeds
    
    try:
        if isinstance(log_data, (list, tuple)):
            lines = log_data
        else:
            try:
                lines = log_data.readlines()
            except AttributeError:
                lines = str(log_data).splitlines()
        
        for line in lines:
            try:
                if isinstance(line, bytes):
                    line = line.decode('utf-8', errors='ignore')
                elif not isinstance(line, str):
                    line = str(line)
                
                line = line.strip()
                if not line:
                    continue
                
                # Look for potential seeds in every non-empty line
                find_potential_seeds(line, filename, root)
                
                # Skip unimportant log entries for regular seed processing
                if not is_meaningful_log(line):
                    continue
                
                # Regular seed processing continues as before...
                if any(term in line for term in ['seed', 'Seed', '/seed']):
                    for pattern in seed_patterns:
                        match = pattern.search(line)
                        if match:
                            seed_value = match.group(1)
                            if is_valid_seed(seed_value) and seed_value not in ignored_seeds:
                                # Add to log results
                                ws_log[f'A{row_log}'] = filename
                                ws_log[f'B{row_log}'] = root
                                ws_log[f'C{row_log}'] = line
                                ws_log[f'D{row_log}'] = seed_value
                                row_log += 1
                                
                                # Update unique seeds dictionary
                                if seed_value not in unique_seeds:
                                    unique_seeds[seed_value] = {
                                        'filename': filename,
                                        'world_name': 'Unknown',
                                        'game_mode': 'Unknown',
                                        'generator': 'Unknown',
                                        'version': 'Unknown',
                                        'last_played': 'Unknown',
                                        'path': root,
                                        'times_found': 1
                                    }
                                else:
                                    unique_seeds[seed_value]['times_found'] += 1
                                break
            except:
                continue
    except:
        pass

def process_regular_file_for_logs(file_path, root, filename):
    """Process a regular file for log content"""
    timeout = get_timeout_for_size(file_path)
    @timeout_handler(timeout)
    def read_and_process_file():
        try:
            # Try reading as text first
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                if content:
                    process_log_content(content.splitlines(), filename, root)
                    return True
        except:
            try:
                # If text reading fails, check if binary
                with open(file_path, 'rb') as f:
                    content = f.read()
                    if not is_binary_content(content):
                        # Try processing as text if not binary
                        text_content = content.decode('utf-8', errors='ignore')
                        process_log_content(text_content.splitlines(), filename, root)
                        return True
            except:
                pass
        return False
    
    try:
        if not read_and_process_file():
            return  # Skip file if binary or empty
    except TimeoutException:
        print(f"Skipping {filename} (timeout after {timeout:.1f}s)")
        errors_encountered += 1
        
        ws_errors[f'A{row_errors}'] = filename
        ws_errors[f'B{row_errors}'] = f"Operation timed out (>{timeout:.1f} seconds)"
        ws_errors[f'C{row_errors}'] = root
        row_errors += 1
        
        ws_data[f'A{row_data}'] = filename
        ws_data[f'B{row_data}'] = "Error: Operation timed out"
        ws_data[f'H{row_data}'] = root
        ws_data[f'P{row_data}'] = "Yes"
        ws_data[f'P{row_data}'].fill = error_fill
        row_data += 1
    except Exception as e:
        errors_encountered += 1
        error_msg = str(e)
        traceback_str = traceback.format_exc()
        
        try:
            with open(file_path, 'rb') as f:
                partial_data = f.read()
                if len(partial_data) > 0:
                    ws_corrupted[f'A{row_corrupted}'] = filename
                    ws_corrupted[f'B{row_corrupted}'] = root
                    ws_corrupted[f'C{row_corrupted}'] = "Yes"
                    ws_corrupted[f'D{row_corrupted}'] = error_msg
                    row_corrupted += 1
                    corrupted_files += 1
        except:
            pass
        
        ws_errors[f'A{row_errors}'] = filename
        ws_errors[f'B{row_errors}'] = error_msg
        ws_errors[f'C{row_errors}'] = root
        ws_errors[f'D{row_errors}'] = traceback_str
        row_errors += 1
        
        ws_data[f'A{row_data}'] = filename
        ws_data[f'B{row_data}'] = f"Error: {error_msg}"
        ws_data[f'H{row_data}'] = root
        ws_data[f'P{row_data}'] = "Yes"
        ws_data[f'P{row_data}'].fill = error_fill
        row_data += 1

def process_gz_file(file_path, root, filename):
    """Process a gzipped file for log content"""
    timeout = get_timeout_for_size(file_path)
    @timeout_handler(timeout)
    def read_and_process_gz():
        try:
            # Try to read the gzipped content
            with gzip.open(file_path, 'rb') as f:
                content = f.read(1024)  # Read first 1KB to check
                if is_binary_content(content):
                    return False
                
                # If not binary, read the full content
                f.seek(0)
                content = f.read().decode('utf-8', errors='ignore')
                if content:
                    process_log_content(content.splitlines(), filename, root)
                    return True
        except:
            pass
        return False

    try:
        if not read_and_process_gz():
            return  # Skip file if binary or empty
    except TimeoutException:
        print(f"Skipping {filename} (timeout after {timeout:.1f}s)")
        errors_encountered += 1
        
        ws_errors[f'A{row_errors}'] = filename
        ws_errors[f'B{row_errors}'] = f"Operation timed out (>{timeout:.1f} seconds)"
        ws_errors[f'C{row_errors}'] = root
        row_errors += 1
        
        ws_data[f'A{row_data}'] = filename
        ws_data[f'B{row_data}'] = "Error: Operation timed out"
        ws_data[f'H{row_data}'] = root
        ws_data[f'P{row_data}'] = "Yes"
        ws_data[f'P{row_data}'].fill = error_fill
        row_data += 1
    except Exception as e:
        errors_encountered += 1
        error_msg = str(e)
        traceback_str = traceback.format_exc()
        
        try:
            with open(file_path, 'rb') as f:
                partial_data = f.read()
                if len(partial_data) > 0:
                    ws_corrupted[f'A{row_corrupted}'] = filename
                    ws_corrupted[f'B{row_corrupted}'] = root
                    ws_corrupted[f'C{row_corrupted}'] = "Yes"
                    ws_corrupted[f'D{row_corrupted}'] = error_msg
                    row_corrupted += 1
                    corrupted_files += 1
        except:
            pass
        
        ws_errors[f'A{row_errors}'] = filename
        ws_errors[f'B{row_errors}'] = error_msg
        ws_errors[f'C{row_errors}'] = root
        ws_errors[f'D{row_errors}'] = traceback_str
        row_errors += 1
        
        ws_data[f'A{row_data}'] = filename
        ws_data[f'B{row_data}'] = f"Error: {error_msg}"
        ws_data[f'H{row_data}'] = root
        ws_data[f'P{row_data}'] = "Yes"
        ws_data[f'P{row_data}'].fill = error_fill
        row_data += 1

def update_unique_seed_info(seed, info):
    """Update unique seed information with the most complete data available"""
    if seed not in unique_seeds:
        unique_seeds[seed] = {
            'filename': info.get('filename', 'Unknown'),
            'world_name': info.get('world_name', 'Unknown'),
            'game_mode': info.get('game_mode', 'Unknown'),
            'generator': info.get('generator', 'Unknown'),
            'version': info.get('version', 'Unknown'),
            'last_played': info.get('last_played', 'Unknown'),
            'path': info.get('path', 'Unknown'),
            'times_found': 1,
            'total_time': info.get('total_time', 'Unknown'),
            'spawn_location': info.get('spawn_location', 'Unknown'),
            'data_version': info.get('data_version', 'Unknown'),
            'difficulty': info.get('difficulty', 'Unknown'),
            'hardcore': info.get('hardcore', 'Unknown'),
            'allow_commands': info.get('allow_commands', 'Unknown'),
            'size_on_disk': info.get('size_on_disk', 'Unknown')
        }
    else:
        unique_seeds[seed]['times_found'] += 1
        
        # Update each field if the new data is more complete
        for field in unique_seeds[seed]:
            if field == 'times_found':
                continue
            
            current_value = unique_seeds[seed][field]
            new_value = info.get(field, 'Unknown')
            
            # Keep the more informative value
            if current_value == 'Unknown' and new_value != 'Unknown':
                unique_seeds[seed][field] = new_value
            elif current_value != 'Unknown' and new_value != 'Unknown':
                # For timestamps, keep the most recent
                if field == 'last_played':
                    try:
                        current_time = datetime.strptime(current_value, '%Y-%m-%d %H:%M:%S')
                        new_time = datetime.strptime(new_value, '%Y-%m-%d %H:%M:%S')
                        if new_time > current_time:
                            unique_seeds[seed][field] = new_value
                    except:
                        pass
                # For version numbers, keep the highest
                elif field == 'version':
                    if new_value > current_value:
                        unique_seeds[seed][field] = new_value

def process_nbt_file(file_path, root, filename):
    """Process an NBT file for world data"""
    global row_data, row_errors, row_corrupted, saved_entries, errors_encountered, corrupted_files
    global ws_data, ws_errors, ws_corrupted, ws_all_seeds, row_all_seeds, error_fill, unique_seeds
    
    timeout = get_timeout_for_size(file_path)
    @timeout_handler(timeout)
    def read_and_process_nbt():
        if is_completely_empty(file_path):
            return None
            
        with open(file_path, 'rb') as f:
            header = f.read(3)
            if header.startswith(b'\x1f\x8b'):  # gzip header
                return nbtlib.load(file_path).root.get('Data', {})
            elif header.startswith(b'\x0A'):  # NBT header
                return nbtlib.load(file_path).root.get('Data', {})
            else:
                return None  # Not a valid NBT file
    
    try:
        nbt_data = read_and_process_nbt()
        if nbt_data is None:
            return
        
        seed = find_seed_in_nbt(nbt_data)
        if not seed:
            return
            
        # Get all the world information
        level_name = nbt_data.get('LevelName', 'Unknown')
        game_mode = nbt_data.get('GameType', 'Unknown')
        if isinstance(game_mode, int):
            game_modes = {0: 'Survival', 1: 'Creative', 2: 'Adventure', 3: 'Spectator'}
            game_mode = game_modes.get(game_mode, f'Unknown ({game_mode})')
        
        version = nbt_data.get('Version', {}).get('Name', 'Unknown')
        last_played = nbt_data.get('LastPlayed', 'Unknown')
        if last_played != 'Unknown' and isinstance(last_played, (int, float)):
            from datetime import datetime
            last_played = datetime.fromtimestamp(last_played / 1000.0).strftime('%Y-%m-%d %H:%M:%S')
        
        generator_name = get_generator_name(nbt_data)
        total_time = nbt_data.get('Time', 'Unknown')
        
        spawn_x = nbt_data.get('SpawnX', 'Unknown')
        spawn_y = nbt_data.get('SpawnY', 'Unknown')
        spawn_z = nbt_data.get('SpawnZ', 'Unknown')
        spawn_location = f"X:{spawn_x} Y:{spawn_y} Z:{spawn_z}" if all(coord != 'Unknown' for coord in [spawn_x, spawn_y, spawn_z]) else 'Unknown'
        
        data_version = nbt_data.get('DataVersion', 'Unknown')
        difficulty = nbt_data.get('Difficulty', 'Unknown')
        if isinstance(difficulty, int):
            difficulties = {0: 'Peaceful', 1: 'Easy', 2: 'Normal', 3: 'Hard'}
            difficulty = difficulties.get(difficulty, f'Unknown ({difficulty})')
        
        hardcore = nbt_data.get('hardcore', False)
        allow_commands = nbt_data.get('allowCommands', 'Unknown')
        size_on_disk = nbt_data.get('SizeOnDisk', 'Unknown')
        
        if size_on_disk != 'Unknown' and isinstance(size_on_disk, (int, float)):
            if size_on_disk > 1073741824:
                size_on_disk = f"{size_on_disk/1073741824:.2f} GB"
            elif size_on_disk > 1048576:
                size_on_disk = f"{size_on_disk/1048576:.2f} MB"
            elif size_on_disk > 1024:
                size_on_disk = f"{size_on_disk/1024:.2f} KB"
            else:
                size_on_disk = f"{size_on_disk} bytes"
        
        # Update unique seeds with all available information
        seed_info = {
            'filename': filename,
            'world_name': level_name,
            'game_mode': game_mode,
            'generator': generator_name,
            'version': version,
            'last_played': last_played,
            'path': root,
            'total_time': total_time,
            'spawn_location': spawn_location,
            'data_version': data_version,
            'difficulty': difficulty,
            'hardcore': 'Yes' if hardcore else 'No',
            'allow_commands': 'Yes' if allow_commands else 'No',
            'size_on_disk': size_on_disk
        }
        
        update_unique_seed_info(seed, seed_info)
        
        # Write to Data worksheet
        ws_data[f'A{row_data}'] = filename
        ws_data[f'B{row_data}'] = seed
        ws_data[f'C{row_data}'] = total_time
        ws_data[f'D{row_data}'] = generator_name
        ws_data[f'E{row_data}'] = level_name
        ws_data[f'F{row_data}'] = game_mode
        ws_data[f'G{row_data}'] = spawn_location
        ws_data[f'H{row_data}'] = root
        ws_data[f'I{row_data}'] = version
        ws_data[f'J{row_data}'] = data_version
        ws_data[f'K{row_data}'] = last_played
        ws_data[f'L{row_data}'] = size_on_disk
        ws_data[f'M{row_data}'] = difficulty
        ws_data[f'N{row_data}'] = 'Yes' if hardcore else 'No'
        ws_data[f'O{row_data}'] = 'Yes' if allow_commands else 'No'
        ws_data[f'P{row_data}'] = "No"
        
        row_data += 1
        saved_entries += 1
        
    except TimeoutException:
        print(f"Skipping {filename} (timeout after {timeout:.1f}s)")
        errors_encountered += 1
        
        ws_errors[f'A{row_errors}'] = filename
        ws_errors[f'B{row_errors}'] = f"Operation timed out (>{timeout:.1f} seconds)"
        ws_errors[f'C{row_errors}'] = root
        row_errors += 1
        
        ws_data[f'A{row_data}'] = filename
        ws_data[f'B{row_data}'] = "Error: Operation timed out"
        ws_data[f'H{row_data}'] = root
        ws_data[f'P{row_data}'] = "Yes"
        ws_data[f'P{row_data}'].fill = error_fill
        row_data += 1
    except Exception as e:
        errors_encountered += 1
        error_msg = str(e)
        traceback_str = traceback.format_exc()
        
        try:
            with open(file_path, 'rb') as f:
                partial_data = f.read()
                if len(partial_data) > 0:
                    ws_corrupted[f'A{row_corrupted}'] = filename
                    ws_corrupted[f'B{row_corrupted}'] = root
                    ws_corrupted[f'C{row_corrupted}'] = "Yes"
                    ws_corrupted[f'D{row_corrupted}'] = error_msg
                    row_corrupted += 1
                    corrupted_files += 1
        except:
            pass
        
        ws_errors[f'A{row_errors}'] = filename
        ws_errors[f'B{row_errors}'] = error_msg
        ws_errors[f'C{row_errors}'] = root
        ws_errors[f'D{row_errors}'] = traceback_str
        row_errors += 1
        
        ws_data[f'A{row_data}'] = filename
        ws_data[f'B{row_data}'] = f"Error: {error_msg}"
        ws_data[f'H{row_data}'] = root
        ws_data[f'P{row_data}'] = "Yes"
        ws_data[f'P{row_data}'].fill = error_fill
        row_data += 1

def sanitize_text(text):
    """Sanitize text for Excel by removing or replacing illegal characters"""
    if not isinstance(text, str):
        text = str(text)
    
    # Replace common problematic characters
    text = text.encode('ascii', 'ignore').decode('ascii')
    
    # Remove control characters
    text = ''.join(char for char in text if ord(char) >= 32 or char in '\n\r\t')
    
    # Limit length to avoid Excel cell limits
    return text[:32000] if len(text) > 32000 else text

def write_potential_seeds():
    """Write potential seeds to the Random Strings worksheet"""
    global ws_potential
    
    row = 2
    for number, info in potential_seeds.items():
        try:
            ws_potential[f'A{row}'] = sanitize_text(number)
            ws_potential[f'B{row}'] = sanitize_text(info['filename'])
            ws_potential[f'C{row}'] = sanitize_text(info['context'])
            ws_potential[f'D{row}'] = sanitize_text(info['line'])
            ws_potential[f'E{row}'] = sanitize_text(info['path'])
            ws_potential[f'F{row}'] = sanitize_text(info['confidence'])
            row += 1
        except Exception as e:
            print(f"\nWarning: Could not write row {row} due to invalid characters. Skipping...")
            continue

def write_unique_seeds():
    """Write unique seeds to the All Seeds worksheet"""
    global row_all_seeds, ws_all_seeds, unique_seeds
    
    # Sort seeds by times_found in descending order
    sorted_seeds = sorted(unique_seeds.items(), key=lambda x: x[1]['times_found'], reverse=True)
    
    for seed, info in sorted_seeds:
        try:
            # Write all available information for each unique seed
            ws_all_seeds[f'A{row_all_seeds}'] = sanitize_text(seed)
            ws_all_seeds[f'B{row_all_seeds}'] = sanitize_text(info['filename'])
            ws_all_seeds[f'C{row_all_seeds}'] = sanitize_text(info['world_name'])
            ws_all_seeds[f'D{row_all_seeds}'] = sanitize_text(info['game_mode'])
            ws_all_seeds[f'E{row_all_seeds}'] = sanitize_text(info['generator'])
            ws_all_seeds[f'F{row_all_seeds}'] = sanitize_text(info['version'])
            ws_all_seeds[f'G{row_all_seeds}'] = sanitize_text(info['last_played'])
            ws_all_seeds[f'H{row_all_seeds}'] = sanitize_text(info['path'])
            ws_all_seeds[f'I{row_all_seeds}'] = info['times_found']
            
            # Add additional columns for complete information
            ws_all_seeds[f'J{row_all_seeds}'] = sanitize_text(str(info['total_time']))
            ws_all_seeds[f'K{row_all_seeds}'] = sanitize_text(info['spawn_location'])
            ws_all_seeds[f'L{row_all_seeds}'] = sanitize_text(str(info['data_version']))
            ws_all_seeds[f'M{row_all_seeds}'] = sanitize_text(info['difficulty'])
            ws_all_seeds[f'N{row_all_seeds}'] = sanitize_text(info['hardcore'])
            ws_all_seeds[f'O{row_all_seeds}'] = sanitize_text(info['allow_commands'])
            ws_all_seeds[f'P{row_all_seeds}'] = sanitize_text(info['size_on_disk'])
            
            row_all_seeds += 1
        except Exception as e:
            print(f"\nWarning: Could not write seed {seed} due to invalid characters. Skipping...")
            continue

def should_skip_file(filename):
    """Check if file should be skipped based on filename"""
    skip_files = {
        'raids.dat',
        'raids_end.dat',
        'villages.dat',
        'villages_end.dat',
        'villages_nether.dat',
        'village.dat',  # Added variant
        'scoreboard.dat',
        'capabilities.dat',
        'temple.dat',
        'idcounts.dat',
        'random_sequences.dat',
        'mineshaft.dat',
        'fortress.dat',
        'fortress_index.dat',
        'mansion_index.dat',
        'mineshaft_index.dat',
        'map_0.dat',
        'map_1.dat',
        'map_2.dat',
        'map_3.dat',
        'map_4.dat',
        'map_5.dat',
        'map_6.dat',
        'map_7.dat',
        'map_8.dat'
    }
    
    lower_filename = filename.lower()
    
    # Check exact matches
    if lower_filename in skip_files:
        return True
    
    # Check map files (for any other numbered maps beyond the specific ones)
    if lower_filename.startswith('map_') and lower_filename.endswith('.dat'):
        return True
    
    return False

def main():
    """Main function to run the Minecraft world recovery script"""
    global processed_files, saved_entries, errors_encountered, corrupted_files
    global row_data, row_errors, row_log, row_all_seeds, row_corrupted
    global wb, ws_data, ws_errors, ws_log, ws_all_seeds, ws_corrupted, error_fill, unique_seeds, potential_seeds
    
    print("=== MC World Recovery ===")
    
    # Initialize Excel workbook and worksheets
    initialize_excel_workbook()
    
    # Reset counters and dictionaries
    processed_files = 0
    saved_entries = 0
    errors_encountered = 0
    corrupted_files = 0
    row_data = 2
    row_errors = 2
    row_log = 2
    row_all_seeds = 2
    row_corrupted = 2
    unique_seeds.clear()
    potential_seeds.clear()
    
    minecraft_files = []

    print("\nCollecting files...")
    # Collect files silently
    for root, dirs, files in os.walk(directory_path):
        for filename in files:
            try:
                # Skip files that should not be processed
                if should_skip_file(filename):
                    continue
                    
                file_path = os.path.join(root, filename)
                if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
                    continue
                
                lower_filename = filename.lower()
                if filename.endswith(".dat"):
                    # Try both NBT and gzipped log formats for .dat files
                    minecraft_files.append(("nbt", root, filename, file_path))
                    minecraft_files.append(("gz", root, filename, file_path))
                elif filename.endswith((".log", ".txt")):
                    minecraft_files.append(("log", root, filename, file_path))
                elif filename.endswith(".gz"):
                    minecraft_files.append(("gz", root, filename, file_path))
            except Exception:
                continue
    
    total_files = len(minecraft_files)
    if total_files == 0:
        print("\nNo files found to process!")
        return
        
    print(f"\nScanning {total_files} files...")
    
    last_progress = -1
    processed_paths = set()  # Track which files we've processed to avoid duplicates in output
    
    for idx, (file_type, root, filename, file_path) in enumerate(minecraft_files, 1):
        try:
            # Show progress for every 1%
            progress = int((idx / total_files) * 100)
            if progress > last_progress:
                # Clear the current line and show new progress
                print(f"\rProgress: {progress}% ({idx}/{total_files} files)", end="", flush=True)
                last_progress = progress
            
            # Skip if we've already processed this file in a different format
            if file_path in processed_paths and file_type == "gz":
                continue
                
            processed_files += 1
            processed_paths.add(file_path)
            
            if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
                continue
    
            if file_type == "nbt":
                process_nbt_file(file_path, root, filename)
            elif file_type == "log":
                process_regular_file_for_logs(file_path, root, filename)
            elif file_type == "gz":
                process_gz_file(file_path, root, filename)
        except Exception:
            continue
    
    print("\rProgress: 100% (Complete)")  # Ensure we show 100% at the end
    
    if processed_files == 0:
        print("\nNo files were successfully processed!")
        return
    
    # Write seeds at the end
    print("\nWriting results...")
    write_unique_seeds()
    write_potential_seeds()
    
    while True:
        try:
            output_path = os.path.join(directory_path, "minecraft_worlds_recovery.xlsx")
            wb.save(output_path)
            print("\n=== Complete ===")
            print(f"Files Processed: {processed_files}")
            print(f"Unique Seeds: {len(unique_seeds)}")
            print(f"Random Strings Found: {len(potential_seeds)}")
            print(f"Log Entries: {row_log - 2}")
            print(f"Errors: {errors_encountered}")
            if corrupted_files > 0:
                print(f"Corrupted Files: {corrupted_files}")
            break
        except Exception as e:
            print("\nError saving results. The file might be open in another program.")
            print("Close the file if it's open and try again.")
            retry = input("Try saving again? (y/n): ").lower()
            if retry != 'y':
                print("Results not saved. Exiting...")
                break

if __name__ == '__main__':
    try:
        main()
        input("\nPress Enter to exit...")
    except KeyboardInterrupt:
        print("\nOperation cancelled by user")
    except Exception as e:
        print(f"\nFatal error: {e}")
        traceback.print_exc()
        input("\nPress Enter to exit...")
