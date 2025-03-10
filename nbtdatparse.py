import nbtlib
import os
import gzip
from openpyxl import Workbook, load_workbook

# Set the directory path
directory_path = 'C:/Users/juke32/AppData/Roaming/.minecraft'
directory_path = 'D:/dump'

# Example of ignored seeds
ignored_seeds = [
    "1234567890",  # Example seed to ignore
    "9876543210"   # Another example seed to ignore

]

# Create a new Excel workbook
wb = Workbook()
ws_data = wb.active
ws_data.title = "Data"
ws_errors = wb.create_sheet(title="Errors")
ws_log = wb.create_sheet(title="Log Results")
ws_tried_seeds = wb.create_sheet(title="Tried Seeds")

# Set header row for Data tab
ws_data['A1'] = 'File Name'
ws_data['B1'] = 'Random Seed'
ws_data['C1'] = 'Total Time'
ws_data['D1'] = 'Generator Name'
ws_data['E1'] = 'Level Name'
ws_data['F1'] = 'Game Mode'
ws_data['G1'] = 'Spawn Location'
ws_data['H1'] = 'Path'

# Set header row for Errors tab
ws_errors['A1'] = 'File Name'
ws_errors['B1'] = 'Error Message'
ws_errors['C1'] = 'Path'

# Set header row for Log Results tab
ws_log['A1'] = 'File Name'
ws_log['B1'] = 'Path'
ws_log['C1'] = 'Log Line'

# Set header row for Tried Seeds tab
ws_tried_seeds['A1'] = 'Tried Seeds'

# Adjust column widths
ws_data.column_dimensions['A'].width = 20
ws_data.column_dimensions['B'].width = 20
ws_data.column_dimensions['C'].width = 15
ws_data.column_dimensions['D'].width = 20
ws_data.column_dimensions['E'].width = 20
ws_data.column_dimensions['F'].width = 15
ws_data.column_dimensions['G'].width = 20
ws_data.column_dimensions['H'].width = 50

ws_errors.column_dimensions['A'].width = 20
ws_errors.column_dimensions['B'].width = 50
ws_errors.column_dimensions['C'].width = 50

ws_log.column_dimensions['A'].width = 20
ws_log.column_dimensions['B'].width = 50
ws_log.column_dimensions['C'].width = 100

ws_tried_seeds.column_dimensions['A'].width = 20

# Initialize row counters
row_data = 2
row_errors = 2
row_log = 2
row_tried_seeds = 2

# Initialize counters for processed and saved files
processed_files = 0
saved_entries = 0
errors_encountered = 0

# Load existing tried seeds from Excel if available
try:
    existing_wb = load_workbook(filename=os.path.join(directory_path, "minecraft_worlds.xlsx"))
    existing_ws = existing_wb['Tried Seeds']
    for row in range(2, existing_ws.max_row + 1):
        ignored_seeds.append(str(existing_ws[f'A{row}'].value))  # Ensure it's a string
except Exception as e:
    print(f"Error loading existing tried seeds: {e}")

# Iterate through all .dat files in the directory and its subdirectories
for root, dirs, files in os.walk(directory_path):
    for filename in files:
        if filename.endswith(".dat"):
            file_path = os.path.join(root, filename)
            
            processed_files += 1
            
            try:
                # Attempt to open .dat file as if it were a .gz file
                with gzip.open(file_path, 'rt', encoding='utf-8', errors='ignore') as f:
                    print(f"Successfully opened {filename} as .gz")
                    for line in f:
                        print(line.strip())
            except OSError as e:
                print(f"Error opening {filename} as .gz: {e}")
                
            try:
                # Load the NBT file
                var = nbtlib.load(file_path)
                
                # Extract and write data to Excel
                ws_data[f'A{row_data}'] = filename
                ws_data[f'H{row_data}'] = os.path.dirname(file_path)
                
                # Try to extract each piece of data individually
                try:
                    seed = str(var.root['Data']['RandomSeed'])
                    if seed in ignored_seeds:  # Check if seed is ignored
                        print(f"Ignoring seed {seed} from {filename}")
                        row_data += 1  # Skip this entry if seed is ignored
                        continue
                    
                    ws_data[f'B{row_data}'] = seed
                    
                    # Add seed to ignored seeds list if not already there
                    if seed not in ignored_seeds:
                        ignored_seeds.append(seed)
                        ws_tried_seeds[f'A{row_tried_seeds}'] = seed
                        row_tried_seeds += 1
                except Exception as e:
                    ws_data[f'B{row_data}'] = f"Error: {e}"
                    print(f"Error extracting RandomSeed from {filename}: {e}")
                    
                try:
                    ws_data[f'C{row_data}'] = var.root['Data']['Time']
                except Exception as e:
                    ws_data[f'C{row_data}'] = f"Error: {e}"
                    print(f"Error extracting Time from {filename}: {e}")
                    
                try:
                    ws_data[f'D{row_data}'] = var.root['Data']['generatorName']
                except Exception as e:
                    ws_data[f'D{row_data}'] = f"Error: {e}"
                    print(f"Error extracting generatorName from {filename}: {e}")
                    
                try:
                    ws_data[f'E{row_data}'] = var.root['Data']['LevelName']
                except Exception as e:
                    ws_data[f'E{row_data}'] = f"Error: {e}"
                    print(f"Error extracting LevelName from {filename}: {e}")
                    
                try:
                    if var.root['Data']['GameType'] == 0:
                        ws_data[f'F{row_data}'] = 'Survival'
                    elif var.root['Data']['GameType'] == 1:
                        ws_data[f'F{row_data}'] = 'Creative'
                    elif var.root['Data']['GameType'] == 2:
                        ws_data[f'F{row_data}'] = 'Adventure'
                    elif var.root['Data']['GameType'] == 3:
                        ws_data[f'F{row_data}'] = 'Spectator'
                except Exception as e:
                    ws_data[f'F{row_data}'] = f"Error: {e}"
                    print(f"Error extracting GameType from {filename}: {e}")
                    
                try:
                    spawn_location = f"X={var.root['Data']['SpawnX']}, Y={var.root['Data']['SpawnY']}, Z={var.root['Data']['SpawnZ']}"
                    ws_data[f'G{row_data}'] = spawn_location
                except Exception as e:
                    ws_data[f'G{row_data}'] = f"Error: {e}"
                    print(f"Error extracting Spawn Location from {filename}: {e}")
                
                saved_entries += 1
                row_data += 1
            except ValueError as e:
                ws_errors[f'A{row_errors}'] = filename
                ws_errors[f'B{row_errors}'] = f"ValueError: {e} while processing {filename}"
                ws_errors[f'C{row_errors}'] = os.path.dirname(file_path)
                print(f"ValueError in {filename}: {e}")
                errors_encountered += 1
                row_errors += 1
            except TypeError as e:
                ws_errors[f'A{row_errors}'] = filename
                ws_errors[f'B{row_errors}'] = f"TypeError: {e} while processing {filename}"
                ws_errors[f'C{row_errors}'] = os.path.dirname(file_path)
                print(f"TypeError in {filename}: {e}")
                errors_encountered += 1
                row_errors += 1
            except Exception as e:
                ws_errors[f'A{row_errors}'] = filename
                ws_errors[f'B{row_errors}'] = f"Error: {e} while processing {filename}"
                ws_errors[f'C{row_errors}'] = os.path.dirname(file_path)
                print(f"Error processing {filename}: {e}")
                errors_encountered += 1
                row_errors += 1

# Iterate through all .gz files in the directory and its subdirectories
seed_command_encountered = False
for root, dirs, files in os.walk(directory_path):
    for filename in files:
        if filename.endswith(".gz"):
            file_path = os.path.join(root, filename)
            
            try:
                with gzip.open(file_path, 'rt', encoding='utf-8', errors='ignore') as f:
                    for line in f:
                        if '/seed' in line.lower():
                            seed_command_encountered = True
                        elif seed_command_encountered:
                            if 'seed' in line.lower() and 'level.dat' not in file_path:
                                seed_match = [seed for seed in ignored_seeds if seed in line]
                                if not seed_match:
                                    ws_log[f'A{row_log}'] = filename
                                    ws_log[f'B{row_log}'] = os.path.dirname(file_path)
                                    ws_log[f'C{row_log}'] = line.strip()
                                    row_log += 1
                            seed_command_encountered = False
                        elif 'seed' in line.lower() and 'level.dat' not in file_path:
                            seed_match = [seed for seed in ignored_seeds if seed in line]
                            if not seed_match:
                                ws_log[f'A{row_log}'] = filename
                                ws_log[f'B{row_log}'] = os.path.dirname(file_path)
                                ws_log[f'C{row_log}'] = line.strip()
                                row_log += 1
            except Exception as e:
                print(f"Error reading {filename}: {e}")

# Save the workbook to location
output_dir = directory_path  
if not os.path.exists(output_dir):
    os.makedirs(output_dir)
wb.save(os.path.join(output_dir, "minecraft_worlds.xlsx"))
print("Workbook saved successfully.")

# Print summary
print(f"Total .dat files processed: {processed_files}")
print(f"Entries successfully saved to Excel: {saved_entries}")
print(f"Errors encountered during processing: {errors_encountered}")

# Pause before closing
input("Press Enter to continue...")
